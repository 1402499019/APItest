import openpyxl
import os,time
import excel2img
from openpyxl.styles import PatternFill, colors, Font, Border, Side

#解决获取excel公式结果为None的方案
from win32com.client import Dispatch


class Excel:
    def __init__(self, filepath):
        self.filepath = filepath

    def get_sheet(self, sheetname: str):
        global sheet, wb
        wb = openpyxl.load_workbook(self.filepath)
        sheet = wb[sheetname]  # 等同于wb.get_sheet_by_name(sheetname)
        return sheet

    def get_cell(self, row: int, col: int):
        return sheet.cell(row, col).value


    def just_open(self, filepath):
        """
        解决获取excel公式结果为None的方案
        :param filepath:
        :return:
        """
        xlApp = Dispatch("Excel.Application")
        xlApp.visible = False
        xlBook = xlApp.Workbooks.Open(filepath)
        xlBook.save()
        xlBook.close()

    # 获取excel公式值
    def get_calc_cell(self, sheetname:str, row:int, col:int):
        self.just_open(self.filepath)
        wb2 = openpyxl.load_workbook(self.filepath, data_only=True)
        sheet2 = wb2[sheetname]
        return sheet2.cell(row, col).value

    # 获取行数
    def get_rows(self):
        return sheet.max_row

    # 单元格赋值
    def set_cell(self, sheetname: str, row: int, col: int, value,
                 fontcolor = colors.BLACK,
                 fonttype = "宋体",
                 fontsize = 11,
                 fontbold = False,
                 fontitalic = False,
                 fill = None,
                 hyperlink = False,
                 link = None,
                 border = True
                 ):

        # 打开sheet页
        ws = wb[sheetname]

        # 设定值
        cell = ws.cell(row, col, value)

        # 背景色
        if fill:
            fill = PatternFill(fill_type="solid", fgColor=fill)
            cell.fill = fill

        # 超链接
        if hyperlink:
            if link:
                cell.hyperlink = link

        # 字体样式
        font = Font(color=fontcolor, name=fonttype, size=fontsize, italic=fontitalic, bold=fontbold)
        cell.font = font

        # 设置边框
        if border:
            borderValue = Border(
                left=Side(border_style="thin",color=colors.BLACK),
                right=Side(border_style="thin",color=colors.BLACK),
                top=Side(border_style="thin",color=colors.BLACK),
                bottom=Side(border_style="thin",color=colors.BLACK),
            )
            cell.border = borderValue

        # 保存
        wb.save(self.filepath)

    # 截图，作为正文发送
    def excel_screenshot(self):
        filename = r"C:\Users\zhang.sun\APItest\screenshot" + time.strftime("%Y-%m-%d_%H%M%S") + ".PNG"

        # Excel 路径，保存图片路径，空，截图区域
        excel2img.export_img(self.filepath, filename, "", "模块!A1:J25")
        return filename

